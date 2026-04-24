"""
Indian Stock Market Intelligence API
- Syncs all NSE-listed companies from NSE official CSV
- Fetches detailed financials from Screener.in
- Generates Excel workbooks for download
- Surfaces concall / investor documents from NSE filings
"""
from fastapi import APIRouter, Depends, Query, BackgroundTasks, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime, timezone
import json, re, io, time

from database import get_db, SessionLocal
import models
from auth import get_current_user

router = APIRouter()

# ── Constants ─────────────────────────────────────────────────────────────────
NSE_CSV_URL      = "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv"
SCREENER_BASE    = "https://www.screener.in/company"
NSE_ANNOUNCE_URL = "https://www.nseindia.com/api/corporate-announcements"

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.google.com/",
}

_NSE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.nseindia.com/",
}


# ── Helpers ───────────────────────────────────────────────────────────────────
def _num(s) -> Optional[float]:
    if not s:
        return None
    clean = re.sub(r"[,\s₹%]", "", str(s).strip())
    if clean.startswith("(") and clean.endswith(")"):
        clean = "-" + clean[1:-1]
    try:
        return float(clean)
    except Exception:
        return None


def _extract_table(section) -> dict:
    """Extract a Screener.in financial table as {years:[...], rows:[{label, values}]}."""
    if not section:
        return {}
    table = section.find("table")
    if not table:
        return {}
    headers = [th.get_text(strip=True) for th in table.find_all("th")]
    rows = []
    for tr in table.find_all("tr"):
        cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
        if len(cells) > 1 and cells[0]:
            label = cells[0].replace("+", "").strip()
            rows.append({"label": label, "values": cells[1:]})
    return {"years": headers[1:] if headers else [], "rows": rows}


# ── STEP 1: NSE CSV sync ──────────────────────────────────────────────────────
def sync_nse_csv(db: Session) -> dict:
    """Download NSE equity CSV and upsert all listed companies."""
    import requests, csv, io as _io
    try:
        r = requests.get(NSE_CSV_URL, headers=_HEADERS, timeout=20)
        r.raise_for_status()
        text = r.content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(_io.StringIO(text))
        rows = list(reader)
    except Exception as e:
        return {"error": str(e), "synced": 0}

    synced = 0
    for row in rows:
        symbol = (row.get("SYMBOL") or "").strip()
        name   = (row.get("NAME OF COMPANY") or "").strip()
        if not symbol or not name:
            continue

        series = (row.get(" SERIES") or row.get("SERIES") or "").strip()
        isin   = (row.get(" ISIN NUMBER") or row.get("ISIN NUMBER") or "").strip()
        fv     = _num(row.get(" FACE VALUE") or row.get("FACE VALUE") or "")
        dol    = (row.get(" DATE OF LISTING") or row.get("DATE OF LISTING") or "").strip()

        existing = db.query(models.NSEListing).filter(
            models.NSEListing.symbol == symbol
        ).first()
        if existing:
            existing.name   = name
            existing.isin   = isin
            existing.series = series
            if fv:  existing.face_value      = fv
            if dol: existing.date_of_listing = dol
        else:
            db.add(models.NSEListing(
                symbol=symbol, name=name, series=series,
                isin=isin, face_value=fv, date_of_listing=dol,
            ))
        synced += 1
        if synced % 300 == 0:
            db.commit()

    db.commit()
    total = db.query(models.NSEListing).count()
    return {"synced": synced, "total_in_db": total, "status": "ok"}


# ── STEP 2: Screener.in scraper ───────────────────────────────────────────────
def fetch_screener_company(symbol: str) -> dict:
    """Scrape Screener.in for full company financials."""
    import requests
    from bs4 import BeautifulSoup
    import html as _html

    result = {"symbol": symbol, "success": False}

    for path in ["/consolidated/", "/"]:
        url = f"{SCREENER_BASE}/{symbol}{path}"
        try:
            r = requests.get(url, headers=_HEADERS, timeout=20)
            if r.status_code == 404:
                continue
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")

            # ── Name ──────────────────────────────────────────────────────────
            h1 = soup.find("h1", class_=lambda c: c and "company-name" in c) or soup.find("h1")
            result["name"] = _html.unescape(h1.get_text(strip=True)) if h1 else symbol

            # ── BSE code ──────────────────────────────────────────────────────
            bse_link = soup.find("a", href=lambda h: h and "bseindia.com/stock" in h)
            if bse_link:
                m = re.search(r"/(\d{5,7})/", bse_link.get("href", ""))
                if m:
                    result["bse_code"] = m.group(1)

            # ── Industry / Sector ─────────────────────────────────────────────
            for a in soup.select(".company-links a, .breadcrumb a"):
                href = a.get("href", "")
                txt  = a.get_text(strip=True)
                if "/industry/" in href:
                    result["industry"] = txt
                elif "/sector/" in href:
                    result["sector"] = txt

            # ── Top ratios ────────────────────────────────────────────────────
            ratios_raw = {}
            for li in soup.select("#top-ratios li, .company-ratios li"):
                parts = li.find_all(["span", "b"])
                if len(parts) >= 2:
                    k = parts[0].get_text(strip=True).lower()
                    v = parts[-1].get_text(strip=True)
                    ratios_raw[k] = v

            def _ratio(*keys):
                for k in keys:
                    for rk, rv in ratios_raw.items():
                        if k.lower() in rk:
                            return _num(rv)
                return None

            result["current_price"]  = _ratio("current price", "price")
            result["market_cap"]     = _ratio("market cap", "mkt cap")
            result["pe_ratio"]       = _ratio("stock p/e", "p/e", "pe ratio")
            result["pb_ratio"]       = _ratio("p/b", "price to book", "price/bv")
            result["roce"]           = _ratio("roce")
            result["roe"]            = _ratio("roe", "return on equity")
            result["div_yield"]      = _ratio("div yield", "dividend yield")
            result["debt_equity"]    = _ratio("debt / equity", "d/e", "debt to equity")
            result["eps"]            = _ratio("eps")
            result["book_value"]     = _ratio("book value")
            result["high_52w"]       = _ratio("52w high", "high price")
            result["low_52w"]        = _ratio("52w low", "low price")
            result["ratios_raw"]     = ratios_raw

            # ── P&L ───────────────────────────────────────────────────────────
            pl_sec = soup.find("section", id="profit-loss")
            if pl_sec:
                table_data = _extract_table(pl_sec)
                result["income_statement"] = table_data
                for row in table_data.get("rows", []):
                    lbl  = row["label"].lower()
                    vals = row["values"]
                    if vals:
                        latest = _num(vals[-1])
                        if "sales" in lbl or "revenue" in lbl:
                            result["revenue"] = latest
                        elif "net profit" in lbl or lbl == "pat":
                            result["net_profit"] = latest

            # ── Balance Sheet ─────────────────────────────────────────────────
            bs_sec = soup.find("section", id="balance-sheet")
            if bs_sec:
                result["balance_sheet"] = _extract_table(bs_sec)

            # ── Cash Flow ─────────────────────────────────────────────────────
            cf_sec = soup.find("section", id="cash-flow")
            if cf_sec:
                result["cash_flow"] = _extract_table(cf_sec)

            # ── Quarterly Results ─────────────────────────────────────────────
            q_sec = soup.find("section", id="quarters")
            if q_sec:
                result["quarterly"] = _extract_table(q_sec)

            # ── Shareholding ──────────────────────────────────────────────────
            sh_sec = soup.find("section", id="shareholding")
            if sh_sec:
                sh_data = _extract_table(sh_sec)
                result["shareholding"] = sh_data
                for row in sh_data.get("rows", []):
                    lbl = row["label"].lower()
                    if "promoter" in lbl and row["values"]:
                        result["promoter_holding"] = _num(row["values"][-1])
                        break

            # ── Pros & Cons ───────────────────────────────────────────────────
            pros_el = soup.find("div", class_=lambda c: c and "pros" in (c or "").lower())
            cons_el = soup.find("div", class_=lambda c: c and "cons" in (c or "").lower())
            if pros_el:
                result["pros"] = [li.get_text(strip=True) for li in pros_el.find_all("li")]
            if cons_el:
                result["cons"] = [li.get_text(strip=True) for li in cons_el.find_all("li")]

            # ── Annual Reports ────────────────────────────────────────────────
            docs_sec = soup.find("section", id="documents")
            annual_reports = []
            if docs_sec:
                for a in docs_sec.find_all("a", href=True):
                    href = a["href"]
                    txt  = a.get_text(strip=True)
                    if "annual" in txt.lower() or "bseindia.com/bseplus/AnnualReport" in href:
                        annual_reports.append({"label": txt[:100], "url": href})
            result["annual_reports"] = annual_reports[:10]

            result["success"]     = True
            result["screener_url"] = url
            return result

        except Exception as e:
            result["error"] = str(e)
            continue

    return result


# ── STEP 3: Excel generator ───────────────────────────────────────────────────
def build_excel(symbol: str, company: models.NSEListing, doc_type: str = "all") -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    HDR_FILL    = PatternFill("solid", fgColor="0C4A6E")
    HDR_FONT    = Font(color="FFFFFF", bold=True, size=11)
    SUBHDR_FILL = PatternFill("solid", fgColor="0EA5E9")
    SUBHDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL    = PatternFill("solid", fgColor="F0F9FF")
    NUM_FMT     = '#,##0.00'
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    def add_sheet(title: str, data: dict):
        ws = wb.create_sheet(title=title[:31])
        years = data.get("years", [])
        rows  = data.get("rows", [])
        if not years or not rows:
            ws["A1"] = "No data available"
            return
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(years) + 1)
        c = ws["A1"]
        c.value = f"{company.name} — {title} (₹ Crores)"
        c.font  = Font(bold=True, size=13, color="0C4A6E")
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22
        for col, val in enumerate(["Particulars"] + years, 1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
        ws.row_dimensions[2].height = 18
        for ridx, row in enumerate(rows, start=3):
            lbl = row.get("label", "")
            is_bold = any(k in lbl.lower() for k in ["total", "profit", "net", "operating", "ebitda", "revenue", "sales"])
            lc = ws.cell(row=ridx, column=1, value=lbl)
            lc.font = Font(bold=is_bold, size=10)
            lc.border = thin
            lc.alignment = Alignment(vertical="center", indent=1)
            if is_bold:
                lc.fill = PatternFill("solid", fgColor="E0F2FE")
            for cidx, val in enumerate(row.get("values", []), start=2):
                num = _num(val)
                vc = ws.cell(row=ridx, column=cidx)
                vc.value = num if num is not None else val
                if num is not None:
                    vc.number_format = NUM_FMT
                vc.font = Font(bold=is_bold, size=10)
                vc.border = thin
                vc.alignment = Alignment(horizontal="right", vertical="center")
                if is_bold:
                    vc.fill = PatternFill("solid", fgColor="E0F2FE")
                elif ridx % 2 == 0:
                    vc.fill = ALT_FILL
        ws.column_dimensions["A"].width = 34
        for col in range(2, len(years) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = "B3"

    # Summary sheet
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 22
    ws_sum.merge_cells("A1:B1")
    ws_sum["A1"].value = f"{company.name} ({company.symbol}) — Financial Summary"
    ws_sum["A1"].font  = Font(bold=True, size=14, color="0C4A6E")
    ws_sum["A1"].alignment = Alignment(horizontal="center")
    ws_sum.row_dimensions[1].height = 24
    ws_sum["A2"].value = f"Exchange: NSE  |  ISIN: {company.isin or 'N/A'}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws_sum["A2"].font  = Font(size=9, color="64748B")
    ws_sum.merge_cells("A2:B2")

    sections = [
        ("MARKET DATA", [
            ("Current Price (₹)", company.current_price),
            ("Market Cap (₹ Cr)", company.market_cap),
            ("52W High (₹)", company.high_52w),
            ("52W Low (₹)", company.low_52w),
        ]),
        ("VALUATION", [
            ("P/E Ratio", company.pe_ratio),
            ("P/B Ratio", company.pb_ratio),
            ("EPS (₹)", company.eps),
            ("Book Value (₹)", company.book_value),
            ("Dividend Yield (%)", company.div_yield),
        ]),
        ("PROFITABILITY", [
            ("ROCE (%)", company.roce),
            ("ROE (%)", company.roe),
            ("Revenue TTM (₹ Cr)", company.revenue),
            ("Net Profit TTM (₹ Cr)", company.net_profit),
        ]),
        ("LEVERAGE & OWNERSHIP", [
            ("Debt / Equity", company.debt_equity),
            ("Promoter Holding (%)", company.promoter_holding),
        ]),
        ("COMPANY INFO", [
            ("Industry", company.industry or "N/A"),
            ("Sector", company.sector or "N/A"),
            ("Face Value (₹)", company.face_value),
            ("Date of Listing", company.date_of_listing or "N/A"),
            ("ISIN", company.isin or "N/A"),
            ("BSE Code", company.bse_code or "N/A"),
        ]),
    ]

    ridx = 3
    for section_title, items in sections:
        sh = ws_sum.cell(row=ridx, column=1, value=section_title)
        sh.font = SUBHDR_FONT; sh.fill = SUBHDR_FILL
        sh.alignment = Alignment(horizontal="left", indent=1)
        ws_sum.merge_cells(start_row=ridx, start_column=1, end_row=ridx, end_column=2)
        ridx += 1
        for label, value in items:
            la = ws_sum.cell(row=ridx, column=1, value=label)
            va = ws_sum.cell(row=ridx, column=2, value=value)
            la.font = Font(size=10, color="334155")
            la.border = thin; la.alignment = Alignment(vertical="center", indent=1)
            va.border = thin; va.alignment = Alignment(horizontal="right", vertical="center")
            if isinstance(value, (int, float)):
                va.number_format = NUM_FMT; va.font = Font(size=10, bold=True)
            else:
                va.font = Font(size=10)
            if ridx % 2 == 0:
                la.fill = ALT_FILL; va.fill = ALT_FILL
            ridx += 1
        ridx += 1  # blank row between sections

    sheets_map = {
        "income_statement": "Income Statement",
        "balance_sheet":    "Balance Sheet",
        "cash_flow":        "Cash Flow",
        "quarterly":        "Quarterly Results",
    }
    to_add = {
        "all":              ["income_statement", "balance_sheet", "cash_flow", "quarterly"],
        "income-statement": ["income_statement"],
        "balance-sheet":    ["balance_sheet"],
        "cash-flow":        ["cash_flow"],
        "quarterly":        ["quarterly"],
    }.get(doc_type, ["income_statement", "balance_sheet", "cash_flow"])

    for key in to_add:
        raw = getattr(company, f"{key}_json", None)
        if raw:
            try:
                add_sheet(sheets_map[key], json.loads(raw))
            except Exception:
                pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── STEP 4: NSE Concall fetcher ───────────────────────────────────────────────
def fetch_concall(symbol: str) -> list:
    import requests
    session = requests.Session()
    try:
        session.get("https://www.nseindia.com/", headers=_NSE_HEADERS, timeout=10)
        time.sleep(0.5)
        r = session.get(
            f"{NSE_ANNOUNCE_URL}?index=equities&symbol={symbol.upper()}",
            headers=_NSE_HEADERS, timeout=15
        )
        if r.status_code != 200:
            return []
        data = r.json()
        all_ann = data if isinstance(data, list) else data.get("data", [])
    except Exception:
        return []

    KEYWORDS = [
        "concall", "conference call", "earnings call", "analyst meet",
        "investor meet", "investor presentation", "post result",
        "q1 result", "q2 result", "q3 result", "q4 result",
        "quarterly result", "annual result", "earnings presentation",
        "transcript",
    ]
    found = []
    for ann in all_ann:
        desc     = str(ann.get("desc", "") or ann.get("subject", "") or "")
        txt      = str(ann.get("attchmntText", "") or "")
        combined = (desc + " " + txt).lower()
        file_url = str(ann.get("attchmntFile", "") or "").strip()
        if any(k in combined for k in KEYWORDS) and file_url:
            found.append({
                "title":    desc.strip() or txt[:100],
                "date":     str(ann.get("sort_date", ""))[:10],
                "file_url": file_url,
                "type": (
                    "Concall Transcript"     if any(k in combined for k in ["concall", "conference call", "transcript"]) else
                    "Investor Presentation"  if "presentation" in combined else
                    "Earnings Call"          if "earnings" in combined else
                    "Analyst / Investor Meet"
                ),
            })
            if len(found) >= 15:
                break
    return found


# ── DB save helper ────────────────────────────────────────────────────────────
def _save_screener_to_db(listing: models.NSEListing, data: dict, db: Session):
    for field in ["industry", "sector", "bse_code", "current_price", "market_cap",
                  "pe_ratio", "pb_ratio", "roce", "roe", "debt_equity", "div_yield",
                  "eps", "book_value", "high_52w", "low_52w", "revenue", "net_profit",
                  "promoter_holding"]:
        val = data.get(field)
        if val is not None:
            setattr(listing, field, val)

    for json_field, data_key in [
        ("pros", "pros"), ("cons", "cons"),
        ("income_statement_json", "income_statement"),
        ("balance_sheet_json",    "balance_sheet"),
        ("cash_flow_json",        "cash_flow"),
        ("quarterly_json",        "quarterly"),
        ("annual_reports_json",   "annual_reports"),
    ]:
        val = data.get(data_key)
        if val:
            setattr(listing, json_field, json.dumps(val))

    listing.fetched_at = datetime.now(timezone.utc)
    db.commit()


def _listing_dict(c: models.NSEListing, full=False) -> dict:
    d = {
        "id":              c.id,
        "symbol":          c.symbol,
        "name":            c.name,
        "isin":            c.isin,
        "series":          c.series,
        "industry":        c.industry,
        "sector":          c.sector,
        "bse_code":        c.bse_code,
        "face_value":      c.face_value,
        "date_of_listing": c.date_of_listing,
        "current_price":   c.current_price,
        "market_cap":      c.market_cap,
        "pe_ratio":        c.pe_ratio,
        "pb_ratio":        c.pb_ratio,
        "roce":            c.roce,
        "roe":             c.roe,
        "debt_equity":     c.debt_equity,
        "div_yield":       c.div_yield,
        "eps":             c.eps,
        "book_value":      c.book_value,
        "high_52w":        c.high_52w,
        "low_52w":         c.low_52w,
        "revenue":         c.revenue,
        "net_profit":      c.net_profit,
        "promoter_holding": c.promoter_holding,
        "fetched_at":      c.fetched_at.isoformat() if c.fetched_at else None,
        "has_financials":  bool(c.income_statement_json),
    }
    if full:
        def _j(s):
            try: return json.loads(s) if s else None
            except: return None
        d["pros"]             = _j(c.pros) or []
        d["cons"]             = _j(c.cons) or []
        d["income_statement"] = _j(c.income_statement_json)
        d["balance_sheet"]    = _j(c.balance_sheet_json)
        d["cash_flow"]        = _j(c.cash_flow_json)
        d["quarterly"]        = _j(c.quarterly_json)
        d["annual_reports"]   = _j(getattr(c, "annual_reports_json", None)) or []
    return d


# ═══════════════════════════════════════════════════════════════════════════════
# Routes
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/sync")
def sync_companies_bg(
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Start NSE sync in background."""
    def _bg():
        bg_db = SessionLocal()
        try:
            result = sync_nse_csv(bg_db)
            print(f"[NSE BG Sync] {result}")
        finally:
            bg_db.close()
    background_tasks.add_task(_bg)
    return {"message": "Sync started in background (~30s)", "status": "started"}


@router.post("/sync/now")
def sync_companies_blocking(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Synchronous NSE sync — returns when complete."""
    result = sync_nse_csv(db)
    return result


# Keep old routes for backward compat
@router.post("/sync-nse")
def sync_nse_compat(background_tasks: BackgroundTasks, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    return sync_companies_bg(background_tasks, db, current_user)

@router.post("/sync-nse/sync")
def sync_nse_sync_compat(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    return sync_companies_blocking(db, current_user)


@router.get("/companies")
def list_companies(
    search:        Optional[str]  = Query(None),
    sector:        Optional[str]  = Query(None),
    industry:      Optional[str]  = Query(None),
    series:        Optional[str]  = Query(None),
    has_financials: Optional[bool] = Query(None),
    page:          int = Query(1, ge=1),
    limit:         int = Query(50, le=200),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    q = db.query(models.NSEListing)

    if search and search.strip():
        pattern = f"%{search.strip()}%"
        q = q.filter(
            models.NSEListing.name.ilike(pattern)   |
            models.NSEListing.symbol.ilike(pattern) |
            models.NSEListing.isin.ilike(pattern)
        )
    if sector:
        q = q.filter(models.NSEListing.sector.ilike(f"%{sector}%"))
    if industry:
        q = q.filter(models.NSEListing.industry.ilike(f"%{industry}%"))
    if series:
        q = q.filter(models.NSEListing.series == series.upper())
    if has_financials is not None:
        if has_financials:
            q = q.filter(models.NSEListing.income_statement_json.isnot(None))
        else:
            q = q.filter(models.NSEListing.income_statement_json.is_(None))

    total = q.count()
    companies = (
        q.order_by(
            models.NSEListing.market_cap.desc().nullslast(),
            models.NSEListing.name.asc()
        )
        .offset((page - 1) * limit)
        .limit(limit)
        .all()
    )

    return {
        "total":     total,
        "page":      page,
        "limit":     limit,
        "pages":     max(1, (total + limit - 1) // limit),
        "companies": [_listing_dict(c) for c in companies],
    }


@router.get("/companies/{symbol}")
def get_company(
    symbol: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    listing = db.query(models.NSEListing).filter(
        models.NSEListing.symbol == symbol.upper()
    ).first()
    if not listing:
        raise HTTPException(404, f"Company '{symbol}' not found. Run sync first.")
    return _listing_dict(listing, full=True)


@router.post("/companies/{symbol}/fetch")
def fetch_company_data(
    symbol: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Fetch / refresh data from Screener.in for this company."""
    listing = db.query(models.NSEListing).filter(
        models.NSEListing.symbol == symbol.upper()
    ).first()
    if not listing:
        listing = models.NSEListing(symbol=symbol.upper(), name=symbol.upper())
        db.add(listing)
        db.commit()
        db.refresh(listing)

    data = fetch_screener_company(symbol.upper())
    if not data.get("success"):
        raise HTTPException(422, f"Could not fetch Screener.in data for {symbol}: {data.get('error', 'Unknown error')}")

    if not listing.name or listing.name == symbol.upper():
        listing.name = data.get("name", symbol.upper())
    _save_screener_to_db(listing, data, db)
    db.refresh(listing)
    return _listing_dict(listing, full=True)


@router.get("/companies/{symbol}/download/{doc_type}")
def download_excel(
    symbol:   str,
    doc_type: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    listing = db.query(models.NSEListing).filter(
        models.NSEListing.symbol == symbol.upper()
    ).first()
    if not listing:
        raise HTTPException(404, f"Company {symbol} not found.")
    if not listing.income_statement_json:
        raise HTTPException(422, "No financial data available. Please fetch data first.")

    buf = build_excel(symbol.upper(), listing, doc_type)
    fname = f"{symbol.upper()}_{doc_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/companies/{symbol}/concall")
def get_concall(
    symbol: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    listing = db.query(models.NSEListing).filter(
        models.NSEListing.symbol == symbol.upper()
    ).first()
    if not listing:
        raise HTTPException(404, f"Company {symbol} not found.")

    # Use cached if recent (< 24h)
    if listing.concall_json and listing.concall_fetched_at:
        age = (datetime.now(timezone.utc) - listing.concall_fetched_at.replace(tzinfo=timezone.utc)).total_seconds()
        if age < 86400:
            return {"concalls": json.loads(listing.concall_json), "cached": True}

    concalls = fetch_concall(symbol.upper())
    listing.concall_json       = json.dumps(concalls)
    listing.concall_fetched_at = datetime.now(timezone.utc)
    db.commit()
    return {"concalls": concalls, "cached": False}


# ── Startup ───────────────────────────────────────────────────────────────────
def init_nse_companies():
    """Ensure table exists and sync NSE companies if DB is empty."""
    from sqlalchemy import text
    db = SessionLocal()
    try:
        # Ensure table exists (handles case where model was added after first run)
        from database import engine, Base
        import models as _models
        Base.metadata.create_all(bind=engine)

        # Runtime column migration for annual_reports_json
        try:
            db.execute(text("SELECT annual_reports_json FROM nse_listings LIMIT 1"))
        except Exception:
            try:
                db.execute(text("ALTER TABLE nse_listings ADD COLUMN annual_reports_json TEXT"))
                db.commit()
                print("✅ Added annual_reports_json column")
            except Exception:
                db.rollback()

        count = db.query(_models.NSEListing).count()
        if count == 0:
            print("📈 NSE table empty — syncing from NSE CSV...")
            result = sync_nse_csv(db)
            print(f"✅ NSE sync: {result.get('synced', 0)} companies loaded")
        else:
            print(f"✅ NSE companies ready: {count} in DB")

    except Exception as e:
        print(f"NSE init error: {e}")
        import traceback; traceback.print_exc()
    finally:
        db.close()
